# clientes/tottus.py
import os
import pandas as pd
import json
import openpyxl
import datetime as dt
from utils.seleccion_archivo import seleccionar_archivo
import sys

def resource_path(relative_path):
    """Obtiene la ruta absoluta, compatible con PyInstaller"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



def clean_text(s):
    if pd.isna(s):
        return ""
    try:
        return s.encode('latin1').decode('utf-8')
    except Exception:
        return s


def formatear_fecha_excel(sheet, df, header_map, columnas_fecha):
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_name, value in zip(df.columns, row):
            if col_name not in columnas_fecha:
                continue

            c_idx = header_map.get(col_name)
            if c_idx is None:
                continue

            cell = sheet.cell(row=r_idx, column=c_idx)
            if pd.notna(value):
                cell.value = value  # ya es string con formato dd-mm-yyyy
                cell.number_format = 'DD-MM-YYYY'  # opcional, para que Excel lo reconozca como fecha
            else:
                cell.value = None


def limpiar_consola():
    os.system('cls' if os.name == 'nt' else 'clear')


def run(df_wms, df_cajas, orden_salida=None):
    if not orden_salida:
        raise ValueError("Se requiere número de orden de salida para ejecutar el proceso.")

    with open(resource_path(os.path.join("data", "client_db.json")), "r", encoding="utf-8") as f:   

        clientes = json.load(f)

    while True:
        limpiar_consola()
        try:
            cliente_sel = clientes[1]
            break
        except (ValueError, IndexError):
            print("❌ Selección inválida. Intente de nuevo.")
            input("Presione Enter para continuar...")

    cod_cliente_input = cliente_sel.get("CodCliente", "")

    df = df_wms.copy()

    columnas_necesarias = [
        'Número OC', 'Tax id proveedor', 'Razón social',
        'Fecha de emisión', 'Fecha fin recepción', 'SKU', 'Unidades compradas'
    ]

    for col in columnas_necesarias:
        if col not in df.columns:
            print(f"❌ Columna '{col}' no encontrada en el archivo CSV.")
            return

    columnas_texto = ['Número OC', 'Tax id proveedor', 'Razón social', 'SKU']
    for col in columnas_texto:
        df[col] = df[col].astype(str).apply(clean_text)

    for fecha_col in ['Fecha de emisión', 'Fecha fin recepción']:
        df[fecha_col] = pd.to_datetime(df[fecha_col], dayfirst=True, errors='coerce')
        df[fecha_col] = df[fecha_col].dt.strftime('%Y-%m-%d 00:00:00')

    nro_referencia = df['Número OC'].dropna().iloc[0] if not df['Número OC'].dropna().empty else ""
    nro_referencia = str(nro_referencia).strip()

    df.rename(columns={
        'Número OC': 'NroOrdenCliente',
        'Tax id proveedor': 'CodCliente',
        'Razón social': 'Nombre Cliente',
        'Fecha de emisión': 'FechaEmision',
        'Fecha fin recepción': 'FechaCompromiso',
        'SKU': 'SKU Item',
        'Unidades compradas': 'CantidadSolicitada'
    }, inplace=True)

    vacios = {
        'Direccion': '',
        'Comuna': '',
        'Ciudad': '',
        'Region': '',
        'Pais': '',
        'Observacion': '',
        'Telefono': '',
        'Email': '',
        'TipoDespacho': '',
        'CodTipoFolio': '',
        'Umedida': '',
        'CrossDocking': '',
        'NroCrossDocking': '',
        'MontoTotal': '',
        'NumeroLote': ''
    }

    for k, v in vacios.items():
        df[k] = v

    df['NroReferencia'] = nro_referencia
    df['NroOrdenSalida'] = orden_salida
    df['CodCliente'] = cod_cliente_input
    df['Nombre Cliente'] = cliente_sel.get("NomCliente", "")
    df['CodSucursal'] = cliente_sel.get("CodSucursal", "")
    df['NomSucursal'] = cliente_sel.get("NomSucursal", "")

    columnas_finales = [
        'NroReferencia', 'NroOrdenCliente', 'CodCliente', 'Nombre Cliente', 'CodSucursal', 'NomSucursal',
        'FechaEmision', 'FechaCompromiso', 'Direccion', 'Comuna', 'Ciudad', 'Region', 'Pais', 'Observacion',
        'Telefono', 'Email', 'TipoDespacho', 'CodTipoFolio', 'SKU Item', 'CantidadSolicitada', 'Umedida',
        'CrossDocking', 'NroCrossDocking', 'MontoTotal', 'NumeroLote', 'NroOrdenSalida'
    ]

    df_final = df[columnas_finales].copy()

    # # Calcular cantidad de bultos (cajas) y total unidades
    # # Se asume que "Unidades dimension logística" está en df_wms y se llama exactamente así
    if 'Unidades dimensión logística' in df.columns:
        # Solo calcula las variables pero no las agregues a df_final
        unidades_logistica = df['Unidades dimensión logística'].astype(float)
        cantidad_solicitada = df_final['CantidadSolicitada'].astype(float)
        bultos = cantidad_solicitada / unidades_logistica
        # Si quieres mostrar o usar estos valores, hazlo aquí, pero no agregues a df_final
        print(f"Cantidad Bultos: {bultos.sum():.0f} - Cantidad Unidades: {cantidad_solicitada.sum():.0f}")
    else:
        print("❌ No se encontró la columna 'Unidades dimensión logística' para calcular bultos.")

    output_path = "output/tottus_resultado.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    type_path = os.path.join("output","JAL", "importar tottus.xlsx")

    if os.path.exists(type_path):
        wb_type = openpyxl.load_workbook(type_path)
        ws_type = wb_type.active

        # Para evitar SettingWithCopyWarning, usa .loc para asignar fechas en df_final
        # Convertir fechas a formato ISO-8601 compatible con SQL Server: 'YYYYMMDDTHH:mm:ss'
        for fecha_col in ['FechaEmision', 'FechaCompromiso']:
            if pd.api.types.is_numeric_dtype(df_final[fecha_col]):
                fechas = pd.TimedeltaIndex(df_final[fecha_col], unit='d') + dt.datetime(1899, 12, 30)
            else:
                fechas = pd.to_datetime(df_final[fecha_col], errors='coerce')

            df_final.loc[:, fecha_col] = fechas

            # Formatear solo fecha sin tiempo ni 'T'
            df_final.loc[:, fecha_col] = fechas.dt.strftime('%Y%m%d')

        # Luego al escribir en Excel, escribe el string tal cual
        ws_type.delete_rows(2, ws_type.max_row)

        header_map = {}
        for col_idx in range(1, ws_type.max_column + 1):
            header = ws_type.cell(row=1, column=col_idx).value
            if header:
                header_map[header] = col_idx

        for r_idx, row in enumerate(df_final.itertuples(index=False), start=2):
            for col_name, value in zip(df_final.columns, row):
                c_idx = header_map.get(col_name)
                if c_idx is None:
                    continue

                cell = ws_type.cell(row=r_idx, column=c_idx)

                if col_name in ['FechaEmision', 'FechaCompromiso']:
                    if pd.notna(value):
                        cell.value = value  # string YYYYMMDD
                        cell.number_format = 'General'  # evitar que Excel lo convierta a número
                    else:
                        cell.value = None
                else:
                    cell.value = value

        wb_type.save(type_path)
        print(f"✅ Archivo importar Tottus Generado Correctamente.")
        
    else:
        print(f"❌ No se encontró el archivo de formato {type_path}. Se genera sin formato especial.")
        df_final.to_excel(type_path, index=False, sheet_name='Sheet1')
        print(f"✅ Archivo generado en: {type_path}")
    
    output_path = "output/tottus_resultado.xlsx"
    df_final.to_excel(output_path, index=False, sheet_name='Sheet1')
    print(f"✅ Archivo final generado en: {output_path}")
    return output_path
