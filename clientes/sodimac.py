# clientes/sodimac.py
import os
import pandas as pd
import json
import openpyxl
import datetime as dt
import sys


def resource_path(relative_path):
    """Obtiene la ruta absoluta, compatible con PyInstaller"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def clean_text(s):
    if pd.isna(s):
        return ""
    try:
        return s.encode('latin1').decode('utf-8')
    except Exception:
        return s


def run(df_wms, df_cajas, orden_salida):
    # Cargar clientes desde archivo JSON
    clientes_path = resource_path(os.path.join("data", "client_db.json"))
    if not os.path.exists(clientes_path):
        raise FileNotFoundError(f"No se encontró el archivo {clientes_path}")

    with open(clientes_path, "r", encoding="utf-8") as f:
        clientes = json.load(f)

    # Seleccionar cliente Sodimac (índice 2)
    try:
        cliente_sel = clientes[2]
    except (ValueError, IndexError):
        raise ValueError("Cliente Sodimac no encontrado en client_db.json")

    cod_cliente_input = cliente_sel.get("CodCliente", "")

    df = df_wms.copy()

    columnas_necesarias = [
        'Número OC', 'Tax id proveedor', 'Razón social',
        'Fecha de emisión', 'Fecha fin recepción', 'SKU', 'Unidades compradas'
    ]

    for col in columnas_necesarias:
        if col not in df.columns:
            raise ValueError(f"Columna '{col}' no encontrada en el archivo WMS.")

    columnas_texto = ['Número OC', 'Tax id proveedor', 'Razón social', 'SKU']
    for col in columnas_texto:
        df[col] = df[col].astype(str).apply(clean_text)

    for fecha_col in ['Fecha de emisión', 'Fecha fin recepción']:
        df[fecha_col] = pd.to_datetime(df[fecha_col], dayfirst=True, errors='coerce')
        df[fecha_col] = df[fecha_col].dt.strftime('%Y-%m-%d 00:00:00')

    nro_referencia = df['Número OC'].dropna().iloc[0] if not df['Número OC'].dropna().empty else ""
    nro_referencia = str(nro_referencia).strip()

    df.rename(columns={
        'Número OC': 'NroOrdenCliente',
        'Tax id proveedor': 'CodCliente',
        'Razón social': 'Nombre Cliente',
        'Fecha de emisión': 'FechaEmision',
        'Fecha fin recepción': 'FechaCompromiso',
        'SKU': 'SKU Item',
        'Unidades compradas': 'CantidadSolicitada'
    }, inplace=True)

    vacios = {
        'Direccion': '',
        'Comuna': '',
        'Ciudad': '',
        'Region': '',
        'Pais': '',
        'Observacion': '',
        'Telefono': '',
        'Email': '',
        'TipoDespacho': '',
        'CodTipoFolio': '',
        'Umedida': '',
        'CrossDocking': '',
        'NroCrossDocking': '',
        'MontoTotal': '',
        'NumeroLote': ''
    }

    for k, v in vacios.items():
        df[k] = v

    df['NroReferencia'] = nro_referencia
    df['NroOrdenSalida'] = orden_salida
    df['CodCliente'] = cod_cliente_input
    df['Nombre Cliente'] = cliente_sel.get("NomCliente", "")
    df['CodSucursal'] = cliente_sel.get("CodSucursal", "")
    df['NomSucursal'] = cliente_sel.get("NomSucursal", "")

    columnas_finales = [
        'NroReferencia', 'NroOrdenCliente', 'CodCliente', 'Nombre Cliente', 'CodSucursal', 'NomSucursal',
        'FechaEmision', 'FechaCompromiso', 'Direccion', 'Comuna', 'Ciudad', 'Region', 'Pais', 'Observacion',
        'Telefono', 'Email', 'TipoDespacho', 'CodTipoFolio', 'SKU Item', 'CantidadSolicitada', 'Umedida',
        'CrossDocking', 'NroCrossDocking', 'MontoTotal', 'NumeroLote', 'NroOrdenSalida'
    ]

    df_final = df[columnas_finales]

    # Calcular cantidad de bultos (cajas) y total unidades
    if 'Unidades dimensión logística' in df.columns:
        df_final.loc[:, 'Unidades dimensión logística'] = df['Unidades dimensión logística'].astype(float)
        df_final.loc[:, 'CantidadSolicitada'] = df_final['CantidadSolicitada'].astype(float)
        df_final.loc[:, 'Bultos'] = df_final['CantidadSolicitada'] / df_final['Unidades dimensión logística']
        # No imprimir en Streamlit
    else:
        # No imprimir en Streamlit
        pass

    output_dir = os.path.join("output", "JAL")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "importar sodimac.xlsx")

    if os.path.exists(output_path):
        wb_type = openpyxl.load_workbook(output_path)
        ws_type = wb_type.active

        # Convertir fechas a formato YYYYMMDD
        for fecha_col in ['FechaEmision', 'FechaCompromiso']:
            if pd.api.types.is_numeric_dtype(df_final[fecha_col]):
                fechas = pd.TimedeltaIndex(df_final[fecha_col], unit='d') + dt.datetime(1899, 12, 30)
            else:
                fechas = pd.to_datetime(df_final[fecha_col], errors='coerce')

            df_final.loc[:, fecha_col] = fechas
            df_final.loc[:, fecha_col] = fechas.dt.strftime('%Y%m%d')

        ws_type.delete_rows(2, ws_type.max_row)

        header_map = {}
        for col_idx in range(1, ws_type.max_column + 1):
            header = ws_type.cell(row=1, column=col_idx).value
            if header:
                header_map[header] = col_idx

        for r_idx, row in enumerate(df_final.itertuples(index=False), start=2):
            for col_name, value in zip(df_final.columns, row):
                c_idx = header_map.get(col_name)
                if c_idx is None:
                    continue

                cell = ws_type.cell(row=r_idx, column=c_idx)

                if col_name in ['FechaEmision', 'FechaCompromiso']:
                    if pd.notna(value):
                        cell.value = value  # string YYYYMMDD
                        cell.number_format = 'General'
                    else:
                        cell.value = None
                else:
                    cell.value = value

        wb_type.save(output_path)
    else:
        # Si no existe plantilla, guardar sin formato especial
        df_final.to_excel(output_path, index=False, sheet_name='Sheet1')

    return output_path
